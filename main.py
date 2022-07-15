from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import WHITE
from time_service import TimeService
from works_service import WorksService
from copy import copy
import os
# 시간 관련 작업을 위해 timeservice class import

time_service = TimeService()

# 네이버 웍스 관련 작업을 위해 WorksService class import
works_service = WorksService()

# 프로그램은 매 달 목요일 오후 6시에 실행하도록 crontab 설정

# 기존에 가져와야하는 파일 명 및 년 설정
year, month, nth_week = time_service.get_this_friday()
target_folder = year + "년"
target_file = '주간업무보고_해외데이터팀({}년{}월{}주).xlsx'.format(year, month, nth_week)

# 엑셀 파일 불러오기
wb = load_workbook(target_file)
ws = wb['주간업무']

# 다음 주 월요일 기준 년도, 주 가져 오기
year, month, nth_week = time_service.get_next_friday()

# 새 타이틀 설정

new_title = "주간 업무 보고({}년 {}월 {}주)".format(year, month, nth_week)
ws.cell(row=1, column=1).value = new_title

# 금주 기간 범위 설정
week_range = time_service.get_next_week()
new_period = ' '.join(["기간: ", week_range])
ws.cell(row=4, column=1).value = new_period

# 작업 붙여넣기
for i in range(7, 45):
    ws.cell(row=i, column=3).value = ws.cell(row=i, column=4).value
    ws.cell(row=i, column=3).fill = copy(ws.cell(row=i, column=4).fill)
    ws.cell(row=i, column=4).value = ""
    ws.cell(row=i, column=4).fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

# 비고란 초기화
ws.cell(row=45, column=2).value = ""

# 엑셀 파일 생성 및 저장
new_file_name = '주간업무보고_해외데이터팀({}년{}월{}주).xlsx'.format(year, month, nth_week)
print(new_file_name)
wb.save(new_file_name)

file = open(new_file_name, "rb")
file_size = os.path.getsize(new_file_name)

# works_service.get_bots()
# works_service.get_users()
#
# 파일 id 및 업로드 링크 받아오기
# response = works_service.getFileUploadLink(new_file_name, file).json()
# print(response)

# print(works_service.getFileDownloadLink(response['fileId']).json())

# works_service.make_room()
file_id = works_service.get_file_id(new_file_name, file)
print(file_id)
# works_service.send_message_to_channel(file_id)